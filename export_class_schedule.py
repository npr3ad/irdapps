"""
export_class_schedule.py — TutorBird class schedule export for a date range.

Run locally:  streamlit run apps/export_class_schedule.py
Deployed at:  Streamlit Community Cloud

Required secrets: APP_PASSWORD, TB_API_KEY

Workflow:
  1. Pick a start/end date.
  2. Click "Fetch Schedule" → preview the appointments below.
  3. Click "Download Excel" to save the grid (Teacher, Appointment Type,
     Date, Time, # Students Scheduled).

TutorBird times are stored as naive Pacific time strings (PDT/PST depending
on season) — the Time column reflects that as-is, no conversion applied.
"""

import io, os
from datetime import date, timedelta
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import streamlit as st

# Load .env for local development (no-op in Streamlit Cloud)
try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent.parent / 'scripts' / '.env')
    load_dotenv(Path(__file__).parent.parent / '.env')
except Exception:
    pass

# ── Credentials ───────────────────────────────────────────────────────────────

def _secret(key: str) -> str:
    """Read from st.secrets first, fall back to os.environ for local dev."""
    try:
        return st.secrets[key]
    except (KeyError, FileNotFoundError, Exception):
        return os.environ.get(key, '')


TB_BASE = 'https://api.tutorbird.com/v1'

def _tb_headers():
    token = _secret('TB_API_KEY')
    return {
        'Authorization': f'Bearer {token}',
        'Accept': 'application/json',
        'Content-Type': 'application/json',
        'x-schoolbox-client-version': '1637',
    }


# ── Auth gate ─────────────────────────────────────────────────────────────────

def _check_password() -> bool:
    if st.session_state.get('_auth'):
        return True
    st.subheader('Sign in')
    pw = st.text_input('Password', type='password', key='_pw')
    if st.button('Sign in', key='_signin'):
        expected = _secret('APP_PASSWORD')
        if pw == expected and expected:
            st.session_state['_auth'] = True
            st.rerun()
        else:
            st.error('Incorrect password.')
    return False


# ── TutorBird API ─────────────────────────────────────────────────────────────

def fetch_events(start_date: date, end_date: date) -> list:
    r = requests.post(
        f'{TB_BASE}/search/calendar/events',
        headers=_tb_headers(),
        json={
            'StartDate': f'{start_date.isoformat()}T00:00:00',
            'EndDate':   f'{end_date.isoformat()}T23:59:59',
            'PageSize':  10000,
            'PageIndex': 0,
        },
        timeout=120,
    )
    r.raise_for_status()
    return r.json().get('ItemSubset', [])


def events_to_rows(events: list) -> list:
    rows = []
    for ev in events:
        teacher      = (ev.get('Teacher') or {}).get('Name', '')
        appt_type    = (ev.get('EventCategory') or {}).get('Name', '')
        start        = ev.get('StartDate', '')
        event_date   = start[:10]
        event_time   = start[11:16]
        num_students = ev.get('CurrentNumberAttending', len(ev.get('Attendances') or []))
        rows.append((teacher, appt_type, event_date, event_time, num_students))
    rows.sort(key=lambda r: (r[2], r[3], r[0].lower()))
    return rows


def build_workbook(rows: list) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Class Schedule'

    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
    alt_fill    = PatternFill('solid', fgColor='D6E4F0')

    headers = ['Teacher', 'Appointment Type', 'Date', 'Time', '# Students Scheduled']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 20

    for i, (teacher, appt_type, d, t, n) in enumerate(rows, start=2):
        fill = alt_fill if i % 2 == 0 else None
        for col, val in enumerate([teacher, appt_type, d, t, n], 1):
            cell = ws.cell(row=i, column=col, value=val)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical='center')

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 20

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:E{len(rows) + 1}'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Streamlit UI ──────────────────────────────────────────────────────────────

st.set_page_config(page_title='IRD Class Schedule', page_icon='🗓️', layout='centered')
st.title('🗓️ Class Schedule Export')

if not _check_password():
    st.stop()

st.caption('Pull TutorBird appointments for a date range: Teacher, Appointment Type, Date, Time, and # Students Scheduled. '
           'Times are Pacific.')

today = date.today()
col1, col2 = st.columns(2)
start_date = col1.date_input('Start date', value=today)
end_date   = col2.date_input('End date', value=today + timedelta(days=6))

if st.button('Fetch Schedule', type='primary'):
    if start_date > end_date:
        st.error('Start date must be on or before end date.')
    else:
        with st.spinner('Fetching from TutorBird…'):
            events = fetch_events(start_date, end_date)
            rows   = events_to_rows(events)
        st.session_state['schedule_rows']  = rows
        st.session_state['schedule_range'] = (start_date, end_date)

if 'schedule_rows' in st.session_state:
    rows = st.session_state['schedule_rows']
    range_start, range_end = st.session_state['schedule_range']

    st.divider()
    st.subheader(f'{len(rows)} appointment(s) — {range_start} to {range_end}')

    if not rows:
        st.warning('No appointments found in this date range.')
    else:
        table = [
            {
                'Teacher':               r[0],
                'Appointment Type':      r[1],
                'Date':                  r[2],
                'Time':                  r[3],
                '# Students Scheduled':  r[4],
            }
            for r in rows
        ]
        st.dataframe(table, use_container_width=True, hide_index=True)

        workbook_buf = build_workbook(rows)
        st.download_button(
            'Download Excel',
            data=workbook_buf,
            file_name=f'class_schedule_{range_start}_to_{range_end}.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
